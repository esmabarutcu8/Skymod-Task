import json
import secrets
import openpyxl
import pandas as pd
from http.server import BaseHTTPRequestHandler, HTTPServer
import urllib.parse
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import io

class RequestHandler(BaseHTTPRequestHandler):

    def do_POST(self):

        if self.path == "/create-certificate":
            try:
                #okunacak verinin uzunluğu
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)

                # Parse JSON data
                data = json.loads(post_data.decode("utf-8"))
                print("Received data:", data)

                # verileri işleme ve sertifika oluşturma
                token = self.generate_unique_token("certificate_info.xlsx")
                name = data["name"]
                lowercase_name = name.lower().replace('ü', 'u').replace('ş', 's').replace('ı', 'i').replace('ö', 'o').replace('ğ', 'g').replace('ç', 'c').replace(' ', '_')


                certificate_data = {
                    "name": name,
                    "educationType": data["educationType"],
                    "duration": data["duration"],
                    "date": data["date"],
                    "organizer": data["organizer"]
                }

                # sertifika pdf oluşturma
                output_pdf = f"./{lowercase_name}.pdf"
                self.write_on_pdf(output_pdf, token, certificate_data)

                #sertifika bağlantısını excel e kaydetme
                link = f"./{lowercase_name}.pdf"
                self.write_data_to_excel(token, link, "certificate_info.xlsx")

                #frontend için başarı mesajı ver sertifika linki oluşturma
                response = {
                    "message": "Sertifika başarıyla oluşturuldu.",
                    "certificate_link": link
                }
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())

            except Exception as e:

                print("Error:", e)
                self.send_response(500)
                self.send_header('Content-type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"message": "Bir hata oluştu!"}).encode())

    def generate_unique_token(self, filename):
        # unique token oluşturma
        try:
            workbook = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        sheet = workbook["Sayfa1"]

        while True:
            token = secrets.token_hex(16)
            if not self.token_exists(sheet, token):
                return token

    def token_exists(self, sheet, token, column_index=2):

        column_letter = sheet.cell(row=1, column=column_index + 1).column_letter
        for cell in sheet[column_letter][1:]:
            if cell.value == token:
                return True
        return False

    def write_data_to_excel(self, token, link, file_name):

        try:
            workbook = openpyxl.load_workbook(file_name)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        sheet = workbook["Sayfa1"]
        sheet.append([token, link])  # Adding new token and link to the Excel sheet

        workbook.save(file_name)

    def write_on_pdf(self, output_pdf, token, data):
        #token ve verileri excele yazdırma
        reader = PdfReader("template_certificate.pdf")
        writer = PdfWriter()

        pdfmetrics.registerFont(TTFont('Poppins-Regular', './font/Poppins/Poppins-Regular.ttf'))
        pdfmetrics.registerFont(TTFont('Poppins-Medium', './font/Poppins/Poppins-Medium.ttf'))

        packet = io.BytesIO()
        can = canvas.Canvas(packet, pagesize=letter)

        can.setFont("Poppins-Medium", 16)
        can.drawString(209, 455, f"{token}")
        can.setFont("Poppins-Medium", 18)

        can.drawString(209, 407, data["name"])
        can.drawString(209, 350, data["educationType"])
        can.drawString(209, 290, data["duration"])
        can.drawString(209, 230, data["date"])
        can.drawString(209, 172, data["organizer"])
        can.save()

        packet.seek(0)
        new_pdf = PdfReader(packet)

        for page_number in range(len(reader.pages)):
            page = reader.pages[page_number]
            page.merge_page(new_pdf.pages[0])
            writer.add_page(page)

        with open(output_pdf, "wb") as output_file:
            writer.write(output_file)

# Starting the HTTP server
def run(server_class=HTTPServer, handler_class=RequestHandler, port=8000):

    server_address = ('', port)
    httpd = server_class(server_address, handler_class)
    print(f"Server running on port {port}...")
    httpd.serve_forever()

if __name__ == '__main__':
    run()
